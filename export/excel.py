"""导出课表到 Excel 文件"""

import asyncio
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def _export_xlsx(courses: List[Dict[str, Any]], filename: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws = wb.active
    if ws is None:
        return
    ws.title = "本学期课表"

    headers = ["课程名称", "任课教师", "学分", "星期", "节次", "周次", "教学楼", "教室"]
    ws.append(headers)

    header_style = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_style
        cell.alignment = align_center
        ws.column_dimensions[get_column_letter(col)].width = 16

    WEEKDAY = {
        1: "周一",
        2: "周二",
        3: "周三",
        4: "周四",
        5: "周五",
        6: "周六",
        7: "周日",
    }

    for item in courses:
        day_raw = item.get("day")
        day = WEEKDAY.get(day_raw, "") if isinstance(day_raw, int) else ""

        start = item.get("start_session")
        duration = item.get("duration")
        section = (
            f"{start}-{start + duration - 1}节"
            if isinstance(start, int) and isinstance(duration, int)
            else ""
        )

        ws.append(
            [
                item.get("course_name", ""),
                (item.get("teacher") or "").replace("*", "").strip(),
                item.get("credit", ""),
                day,
                section,
                item.get("week_desc") or item.get("weeks") or "",
                item.get("building") or item.get("teachingBuildingName") or "",
                item.get("classroom") or item.get("classroomName") or "",
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = align_center

    wb.save(filename)


async def export_timetable_excel(courses: List[Dict[str, Any]], filename: str) -> None:
    """导出课表到 Excel"""
    await asyncio.to_thread(_export_xlsx, courses, filename)

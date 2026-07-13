"""课表明细导出"""

import asyncio
from collections.abc import Mapping, Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = (
    "课程名称",
    "任课教师",
    "学分",
    "星期",
    "节次",
    "周次",
    "教学楼",
    "教室",
)
WEEKDAY_NAMES = {
    1: "周一",
    2: "周二",
    3: "周三",
    4: "周四",
    5: "周五",
    6: "周六",
    7: "周日",
}
COLUMN_WIDTHS = (24, 16, 10, 10, 14, 20, 18, 18)
UNKNOWN_SORT_POSITION = 99


def _as_int(value: object) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, str):
        try:
            return int(value)
        except ValueError:
            return None
    return None


def _as_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _course_sort_key(item: Mapping[str, object]) -> tuple[int, int, str]:
    day = _as_int(item.get("day"))
    start_session = _as_int(item.get("start_session"))
    return (
        day if day is not None else UNKNOWN_SORT_POSITION,
        start_session if start_session is not None else UNKNOWN_SORT_POSITION,
        _as_text(item.get("course_name")),
    )


def _format_section(start: object, duration: object) -> str:
    start_number = _as_int(start)
    duration_number = _as_int(duration)
    if start_number is None or duration_number is None or duration_number < 1:
        return ""
    end_number = start_number + duration_number - 1
    return f"{start_number}-{end_number}节"


def _row_values(item: Mapping[str, object]) -> list[object]:
    day = _as_int(item.get("day"))
    weekday = WEEKDAY_NAMES.get(day, "") if day is not None else ""
    teacher = _as_text(item.get("teacher")).replace("*", "").strip()
    week_description = item.get("week_desc") or item.get("weeks") or ""
    building = item.get("building") or item.get("teachingBuildingName") or ""
    classroom = item.get("classroom") or item.get("classroomName") or ""
    return [
        _as_text(item.get("course_name")),
        teacher,
        item.get("credit", ""),
        weekday,
        _format_section(item.get("start_session"), item.get("duration")),
        _as_text(week_description),
        _as_text(building),
        _as_text(classroom),
    ]


def _export_xlsx(
    courses: Sequence[Mapping[str, object]],
    filename: Path,
) -> Path:
    filename.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    try:
        worksheet = workbook.active
        if worksheet is None:
            msg = "workbook did not create an active worksheet"
            raise RuntimeError(msg)

        worksheet.title = "本学期课表"
        worksheet.freeze_panes = "A2"
        worksheet.append(HEADERS)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        centered = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        for column, width in enumerate(COLUMN_WIDTHS, start=1):
            cell = worksheet.cell(row=1, column=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered
            worksheet.column_dimensions[get_column_letter(column)].width = width

        for item in sorted(courses, key=_course_sort_key):
            worksheet.append(_row_values(item))

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = centered

        worksheet.auto_filter.ref = worksheet.dimensions
        workbook.save(filename)
    finally:
        workbook.close()
    return filename


async def export_timetable_excel(
    courses: Sequence[Mapping[str, object]],
    filename: str | Path,
) -> Path:
    """导出课表"""
    return await asyncio.to_thread(_export_xlsx, courses, Path(filename))
